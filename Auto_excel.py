# The purpose of this script is to explore python
# The goal is to develop a program that can scrape an xlsx file for
# charge hours on projects to track available funds throughtout
# the fiscal year automatically.
# Matthew Kelly 10-04-2020

# Testing github integration with Atom!

import pandas as pd # panda reads/manipulates xlsx files
from openpyxl import load_workbook # openpyxl reads/manipulates xlsx files
import tkinter as tk # GUI package


#############################################################
# Functions
#############################################################

# This function will return a dictionary
# with a dataframe for each worksheet in the workbook
# Loop through the sheets and make a dataframe array that holds each one
# This code imports the output worksheets as a dictionary acting like a dataframe array.
# The indices of the dictionary that designate the sheets are the sheet names themselves.
def xl2df(filename):
    wb = load_workbook(filename) # Load the excel sheet with openpyxl
    ws = wb.active
    _sheets = wb.sheetnames # List of sheet names within the workbook
    _df_sheet_collection = {} # A dictionary to hold all the worksheets
    for i in range(len(_sheets)): # Loop through all of the sheets
        ws = wb[_sheets[i]]
        _data = ws.values
        _cols = next(_data)[0:]
        _data = list(_data)
        _idx = [r[0] for r in _data]
        _df_sheet_collection[_sheets[i]] = pd.DataFrame(_data, index=_idx, columns=_cols)

    return _df_sheet_collection

# This function will return a list of
# positions where element exists
# in the dataframe.
def getIndexes(dfObj, value):

    # Empty list
    listOfPos = []

    # isin() method will return a dataframe with
    # boolean values, True at the positions
    # where element exists
    result = dfObj.isin([value])

    # any() method will return
    # a boolean series
    seriesObj = result.any()

    # Get list of column names where
    # element exists
    columnNames = list(seriesObj[seriesObj == True].index)

    # Iterate over the list of columns and
    # extract the row index where element exists
    for col in columnNames:
        rows = list(result[col][result[col] == True].index)

        for row in rows:
            listOfPos.append((row, col))

    # This list contains a list tuples with
    # the index of element in the dataframe
    return listOfPos

# Use 3 decimal places in output display
pd.set_option("display.precision", 3)
# Don't wrap repr(DataFrame) across additional lines
pd.set_option("display.expand_frame_repr", False)
# Set max rows displayed in output to 25
pd.set_option("display.max_rows", 25)


#############################################################
# # Load workbooks
#############################################################

input_file_name = 'payroll/payperiod_report.xlsx'
_df_input_wb = xl2df(input_file_name)
_sheets_in = list(_df_input_wb)
_employee_in = list(_df_input_wb[_sheets_in[0]].Employee) # Lists all employees
_employee_in = list(dict.fromkeys(_employee_in)) # Eliminate duplicate entries
_project_in = list(_df_input_wb[_sheets_in[0]].JON) # Lists all employees
_project_in = list(dict.fromkeys(_project_in)) # Eliminate duplicate entries

output_file_name = 'payroll/test.xlsx'
_df_output_wb = xl2df(output_file_name)
_sheets_out = list(_df_output_wb)
_employee_out = list(_df_output_wb[_sheets_out[0]].index) # List all employees
_employee_out = list(dict.fromkeys(_employee_out)) # Eliminate duplicate entries
_project_out = list(_df_output_wb) # List all projects

# Load the output file for updating
wb = load_workbook(output_file_name)


#############################################################
# Data processing
#############################################################

charge_hours = _df_input_wb[_sheets_in[0]].groupby(["JON", "Employee"])["Hours"].sum().reset_index() # Sum hours of an employee on each project
charge_hours = charge_hours.dropna().reset_index() # Eliminate rows where employee had no charged hours
sum_hours_charged = pd.DataFrame(charge_hours) # Converts the object to a dataframe

# Construct a charge hours data DataFrame
df_charge_hours = pd.DataFrame(index = _employee_out) # Initialize the DataFrame with the rows and cloumns projects and employees that were active in the pay period
# Populte DataFrame with payperiod charge hours
# This DataFrame will consist of rows of employees and columns of project job order numbers (JON)
_pay_period_end_date = 'PP3'
dict_charge_hours = {}
for i in range(sum_hours_charged.shape[0]):
    _row = sum_hours_charged.Employee[i]
    _col = sum_hours_charged.JON[i]
    # The values represent hours worked by each employee for each JON for the pay period
    df_charge_hours.loc[_row,_col] = sum_hours_charged.Hours[i] # The values represent hours worked by each employee for each JON for the pay period
    # Dictionary first key = worksheet title, second key = pay period end date, values = total hours worked by each employee
    # structured so each set of keys have a dataframe equal to what needs to be updated in the excel sheet
    _temp_sheet = sum_hours_charged.JON[i]
    dict_charge_hours[str(_temp_sheet), _pay_period_end_date] = df_charge_hours.loc[:,_col]

print('Input file successfully uploaded and processed....' + input_file_name)

# Serch for the row number of each employee
_employee_row = list(range(len(_employee_in)))
for i in range(len(_employee_in)): # Loop through employees
    _employee_row[i] = getIndexes(_df_output_wb[_sheets_out[0]], _employee_in[i])

# Serch for the col number of the current pay period
#_pay_period_col = getIndexes(_df_output_wb[_sheets_out[0]], _pay_period_end_date)
for i in range(len(_df_output_wb[_sheets_out[0]].columns)):
    if _pay_period_end_date == _df_output_wb[_sheets_out[0]].columns[i]:
        _pay_period_col = i + 1 # Added 1 ro account for the row index in the excel file


#############################################################
# Update output file
#############################################################

# There may be a more elegant solution than just adding constants onto the counters for rows and columns
# Using the openpyxl library may help clean this section of code up a bit
for i in range(len(_sheets_out)):
    for j in range(df_charge_hours.shape[1]): # Loop through employees
        _row = _employee_in[j]
        _col = _project_in[i]
        _value = df_charge_hours.loc[_row,_col]
        wb[_sheets_out[i]].cell(row = j+2, column = _pay_period_col, value = _value) # Alter excel sheet

wb.save(output_file_name)


#############################################################
# Output
#############################################################

print('Output file successfully updated....' + output_file_name)
for i in _sheets_out:
    print(' ')
    print('Project: '+i)
    print('Pay period: '+str(_pay_period_end_date))
    dict_charge_hours[i, _pay_period_end_date] = dict_charge_hours[i, _pay_period_end_date].dropna() # Eliminate rows where employee had no charged hours
    print(dict_charge_hours[i, _pay_period_end_date])


#############################################################
# Old code
#############################################################
'''
dtypes = {
    "Employee": "category",
    "Date": "category",
    "Hours": "float64",
    "JON": "category",
}

input_file_name = 'payroll/payperiod_report.xlsx'
df_input = pd.read_excel(
    "payroll/payperiod_report.xlsx",
    dtype = dtypes,
    usecols = list(dtypes)
)

dtypes = {
    "Employee": "string",
    "PP1": "category",
    "PP2": "category",
    "PP3": "category",
    "PP4": "category",
    "Total": "category",
}

output_file_name = 'payroll/test.xlsx'
df_output = pd.read_excel(
    "payroll/test.xlsx",
    dtype = dtypes,
    usecols = list(dtypes)
)

#############################################################
# Data processing
#############################################################

# This section creates a list of job order numbers and employees that were active during the payperiod with no duplicates for referance
employee = df_input['Employee'].tolist() # Lists all employees
employee = list(dict.fromkeys(employee)) # Eliminate duplicate entries
project = df_input['JON'].tolist() # Lists all project job order numbers (JON)
project = list(dict.fromkeys(project)) # Eliminate duplicate entries
sheets = wb.sheetnames # Lists all sheets in the output file


# Sums up all of the charge hours each individual has for a project during the entire payperiod
# The result is a list of cumulative hours charged on each project by an employee
charge_hours = df_input.groupby(["JON", "Employee"])["Hours"].sum().reset_index() # Sum hours of an employee on each project
charge_hours = charge_hours.dropna().reset_index() # Eliminate rows where employee had no charged hours
sum_hours_charged = pd.DataFrame(charge_hours) # Converts the object to a dataframe
print(sum_hours_charged.Employee)
# Construct a charge hours data DataFrame
df_charge_hours = pd.DataFrame(index = employee, columns = project) # Initialize the DataFrame with the rows and cloumns projects and employees that were active in the pay period
# Populte DataFrame with payperiod charge hours
# This DataFrame will consist of rows of employees and columns of project job order numbers (JON)
# The values represent hours worked by each employee for each JON for the pay period
for i in range(sum_hours_charged.shape[0]):
    _row = sum_hours_charged.Employee[i]
    _col = sum_hours_charged.JON[i]
    df_charge_hours.loc[_row,_col] = sum_hours_charged.Hours[i]

print(df_charge_hours.index)

print('Input file successfully uploaded and processed....' + input_file_name)

# Serch for the row number of each employee
employee_row = list(range(len(employee)))
for i in range(len(employee)): # Loop through employees
    employee_row[i] = getIndexes(df_output, employee[i])

# Find the column for the payperiod
# Currently this uses a place holder but it neeeds to be changed to take dates as strings
pay_period_end_date = 'PP3'
for i in range(len(df_output.columns)):
    if pay_period_end_date == df_output.columns[i]:
        pay_period_ending = i+1 # Added 1 ro account for the row index in the excel file


#############################################################
# Update output file
#############################################################

# There may be a more elegant solution than just adding constants onto the counters for rows and columns
# Using the openpyxl library may help clean this section of code up a bit
for i in range(len(sheets)):
    for j in range(df_charge_hours.shape[1]): # Loop through employees
        _row = employee[j]
        _col = project[i]
        _value = df_charge_hours.loc[_row,_col]
        wb[sheets[i]].cell(row = j+2, column = pay_period_ending, value = _value) # Alter excel sheet

print('==============')
print(employee)
print(df_charge_hours)


print(df_charge_hours)
# wb.save(output_file_name)
print('Output file successfully updated....' + output_file_name)
'''
